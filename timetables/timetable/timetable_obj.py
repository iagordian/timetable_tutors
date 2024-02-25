from collections import defaultdict
from typing import Optional
import uuid
from functools import total_ordering
from flask import render_template
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import io
from typing import List
from sqlalchemy.orm import Query

class TM_Ceil:

  def __init__(self, val: Optional[uuid.UUID]=None, label: str=''):
    self.val = val
    self.label = label

  def __repr__(self):
    return f'[<{self.val}> {self.label}]'

@total_ordering
class TM_Row:

  def __init__(self, tutor_id: Optional[uuid.UUID], tutor_name: str, lessons: Query):
    self.tutor_id = tutor_id
    self.tutor_name = tutor_name
    self.cursor = 1

    self.lessons = defaultdict(lambda: TM_Ceil())

    if lessons:
      for lesson in lessons:
        self.lessons[lesson.lesson_num] = TM_Ceil(lesson.student_id, lesson.student.student_name)

  def __repr__(self):
    return f'[<{self.tutor_id}> {self.tutor_name}]: {" ".join([str(self.lessons[i]) for i in range(1, 9)])}'

  def __iter__(self):
    return self

  def __next__(self):

    '''Объект позволяет итерироваться по собственным ячейкам'''

    if self.cursor > 8:
      raise StopIteration

    lesson = self.lessons[self.cursor]
    self.cursor += 1
    return lesson

  def __eq__(self, other):

    if isinstance(other, TM_Row):
      return self.tutor_name == other.tutor_name

    return NotImplemented

  def __lt__(self, other):

    if isinstance(other, TM_Row):
      return self.tutor_name < other.tutor_name

    return NotImplemented


class Timetable:

  def __init__(self, lessons: List[dict], students: Query):

    self.rows = []
    for tutor_lessons in lessons:
      self.rows.append(TM_Row(tutor_lessons['tutor'].tutor_id,
                              tutor_lessons['tutor'].tutor_name,
                              lessons = tutor_lessons['lessons']))

    self.options = {student.student_id: student.student_name for student in students if student.active}
    self.options[None] = ''

    self.iteration_size = len(self.rows) - 1
    self.cursor = 0

    self.sort()

  def __iter__(self):
    return self

  def __next__(self):

    '''Объект позволяет итерироваться по собственным строкам'''

    if self.cursor > self.iteration_size:
      raise StopIteration

    row = self.rows[self.cursor]
    self.cursor += 1
    return row

  def __repr__(self):
    return '[' + '\n'.join([str(row) for row in self.rows]) + ']'

  def __call__(self):
    '''Рендер HTML-кода на основании объекта расписания'''
    return render_template('/timetable/timetable.html', timetable=self)

  def sort(self):
    self.rows = sorted(self.rows)

  def excel(self) -> io.BytesIO:

    column_names = [chr(ind + 65) for ind in range(9)]
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Расписание'

    for ind, head in enumerate(['Тьютор', 1, 2, 3, 4, 5, 6, 7, 8], start=1):
      ceil = ws.cell(1, ind)
      ceil.value = head
      ceil.border = border
      ceil.font = Font(name='Times New Roman', size=14, bold=True)
      ceil.alignment = Alignment(horizontal='center')
      ceil.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')

    for row_ind, row in enumerate(self, start=2):
      ceil = ws.cell(row_ind, 1)
      ceil.value = row.tutor_name
      ceil.border = border
      ceil.font = Font(name='Times New Roman', size=12)
      ceil.alignment = Alignment(horizontal='center')
      ceil.fill = PatternFill(start_color='ff7f50', end_color='ff7f50', fill_type='solid')

      for ceil_ind, row_ceil in enumerate(row, start=2):
        ceil = ws.cell(row_ind, ceil_ind)
        ceil.value = row_ceil.label
        ceil.border = border
        ceil.font = Font(name='Times New Roman', size=12)
        ceil.alignment = Alignment(horizontal='center')
        ceil.fill = PatternFill(start_color='ffd700', end_color='ffd700', fill_type='solid')

    for column_name in column_names:
      ws.column_dimensions[column_name].width = 20

    bytes_container = io.BytesIO()
    wb.save(bytes_container)
    bytes_container.seek(0)

    return bytes_container